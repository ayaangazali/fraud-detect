"""
Excel and CSV Export Generator for Phase 7
Handles Excel (xlsx) and CSV file generation
"""
import os
import csv
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Excel file generator with formatting"""
    
    def __init__(self):
        self.wb = Workbook()
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_screening_summary_excel(
        self,
        data: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> str:
        """
        Generate Excel file for screening summary
        
        Args:
            data: Report data
            metadata: Report metadata
            
        Returns:
            File path
        """
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Create Summary sheet
        ws_summary = self.wb.create_sheet("Summary")
        self._add_summary_sheet(ws_summary, data, metadata)
        
        # Create Detailed Data sheet
        ws_details = self.wb.create_sheet("Screening Details")
        self._add_screening_details_sheet(ws_details, data)
        
        # Create Top Matches sheet
        ws_matches = self.wb.create_sheet("Top Blacklist Matches")
        self._add_top_matches_sheet(ws_matches, data.get('top_blacklist_matches', []))
        
        # Create Trend sheet
        ws_trend = self.wb.create_sheet("Trend Analysis")
        self._add_trend_sheet(ws_trend, data.get('screening_trend', []))
        
        # Save file
        filename = f"screening_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        self.wb.save(filepath)
        
        logger.info(f"Excel report generated: {filepath}")
        return filepath
    
    def generate_flagged_items_excel(
        self,
        data: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> str:
        """Generate Excel file for flagged items report"""
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Summary sheet
        ws_summary = self.wb.create_sheet("Summary")
        self._add_flagged_summary_sheet(ws_summary, data, metadata)
        
        # Detailed items sheet
        ws_items = self.wb.create_sheet("Flagged Items")
        self._add_flagged_items_sheet(ws_items, data.get('flagged_items', []))
        
        # Save file
        filename = f"flagged_items_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        self.wb.save(filepath)
        
        logger.info(f"Excel report generated: {filepath}")
        return filepath
    
    def generate_case_history_excel(
        self,
        data: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> str:
        """Generate Excel file for case history report"""
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Summary sheet
        ws_summary = self.wb.create_sheet("Summary")
        self._add_case_summary_sheet(ws_summary, data, metadata)
        
        # Cases details sheet
        ws_cases = self.wb.create_sheet("Cases")
        self._add_cases_sheet(ws_cases, data.get('cases', []))
        
        # Save file
        filename = f"case_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        self.wb.save(filepath)
        
        logger.info(f"Excel report generated: {filepath}")
        return filepath
    
    def generate_compliance_audit_excel(
        self,
        data: Dict[str, Any],
        metadata: Dict[str, Any]
    ) -> str:
        """Generate Excel file for compliance audit report"""
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Summary sheet
        ws_summary = self.wb.create_sheet("Summary")
        self._add_audit_summary_sheet(ws_summary, data, metadata)
        
        # Audit trail sheet
        ws_audit = self.wb.create_sheet("Audit Trail")
        self._add_audit_trail_sheet(ws_audit, data.get('audit_trail', []))
        
        # Save file
        filename = f"compliance_audit_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        self.wb.save(filepath)
        
        logger.info(f"Excel report generated: {filepath}")
        return filepath
    
    # Helper methods for sheet creation
    
    def _add_summary_sheet(self, ws, data, metadata):
        """Add summary sheet with key metrics"""
        # Title
        ws['A1'] = metadata.get('title', 'Screening Summary Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Metadata
        row = 3
        ws[f'A{row}'] = 'Generated:'
        ws[f'B{row}'] = metadata.get('generated_at', '')
        row += 1
        ws[f'A{row}'] = 'Period:'
        ws[f'B{row}'] = f"{metadata.get('date_from', '')} to {metadata.get('date_to', '')}"
        row += 2
        
        # Key Metrics
        ws[f'A{row}'] = 'Key Metrics'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        metrics = [
            ('Total Screenings:', data.get('total_screenings', 0)),
            ('Total Matches:', data.get('total_matches', 0)),
            ('Critical Matches:', data.get('critical_matches', 0)),
            ('High Matches:', data.get('high_matches', 0)),
            ('Medium Matches:', data.get('medium_matches', 0)),
            ('Low Matches:', data.get('low_matches', 0)),
            ('Match Rate:', f"{data.get('match_rate', 0)}%"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _add_screening_details_sheet(self, ws, data):
        """Add screening details with entity breakdown"""
        # Headers
        headers = ['Entity Type', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Data
        entity_breakdown = data.get('entity_breakdown', {})
        total = sum(entity_breakdown.values()) if entity_breakdown else 1
        
        row = 2
        for entity_type, count in entity_breakdown.items():
            ws.cell(row=row, column=1, value=entity_type)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
            row += 1
    
    def _add_top_matches_sheet(self, ws, matches):
        """Add top blacklist matches"""
        headers = ['Blacklist Name', 'Source', 'Match Count', 'Avg Score']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, match in enumerate(matches, 2):
            ws.cell(row=row, column=1, value=match.get('name', ''))
            ws.cell(row=row, column=2, value=match.get('source', ''))
            ws.cell(row=row, column=3, value=match.get('count', 0))
            ws.cell(row=row, column=4, value=f"{match.get('avg_score', 0):.1f}")
    
    def _add_trend_sheet(self, ws, trend):
        """Add trend analysis"""
        headers = ['Date', 'Screenings', 'Matches']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, day in enumerate(trend, 2):
            ws.cell(row=row, column=1, value=day.get('date', ''))
            ws.cell(row=row, column=2, value=day.get('count', 0))
            ws.cell(row=row, column=3, value=day.get('matches', 0))
    
    def _add_flagged_summary_sheet(self, ws, data, metadata):
        """Add flagged items summary"""
        ws['A1'] = metadata.get('title', 'Flagged Items Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        metrics = [
            ('Total Flagged:', data.get('total_flagged', 0)),
            ('Pending:', data.get('pending_count', 0)),
            ('Approved:', data.get('approved_count', 0)),
            ('Rejected:', data.get('rejected_count', 0)),
            ('Resolved:', data.get('resolved_count', 0)),
            ('Avg Resolution Time:', f"{data.get('average_resolution_time', 0)} days"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _add_flagged_items_sheet(self, ws, items):
        """Add flagged items details"""
        headers = ['ID', 'Entity Name', 'Type', 'Blacklist Match', 'Score', 'Status', 'Severity', 'Flagged By', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, item in enumerate(items, 2):
            ws.cell(row=row, column=1, value=item.get('id', ''))
            ws.cell(row=row, column=2, value=item.get('kamco_name', ''))
            ws.cell(row=row, column=3, value=item.get('kamco_type', ''))
            ws.cell(row=row, column=4, value=item.get('blacklist_name', ''))
            ws.cell(row=row, column=5, value=item.get('match_score', ''))
            ws.cell(row=row, column=6, value=item.get('status', ''))
            ws.cell(row=row, column=7, value=item.get('severity', ''))
            ws.cell(row=row, column=8, value=item.get('flagged_by', ''))
            ws.cell(row=row, column=9, value=item.get('created_at', ''))
    
    def _add_case_summary_sheet(self, ws, data, metadata):
        """Add case history summary"""
        ws['A1'] = metadata.get('title', 'Case History Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        metrics = [
            ('Total Cases:', data.get('total_cases', 0)),
            ('Open Cases:', data.get('open_cases', 0)),
            ('Closed Cases:', data.get('closed_cases', 0)),
            ('Approved Cases:', data.get('approved_cases', 0)),
            ('Rejected Cases:', data.get('rejected_cases', 0)),
            ('Avg Resolution Time:', f"{data.get('average_resolution_time', 0)} days"),
            ('SLA Compliance:', f"{data.get('sla_compliance_rate', 0)}%"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _add_cases_sheet(self, ws, cases):
        """Add cases details"""
        headers = ['Case Number', 'Status', 'Priority', 'Created', 'Resolved', 'Resolution Time']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, case in enumerate(cases, 2):
            ws.cell(row=row, column=1, value=case.get('case_number', ''))
            ws.cell(row=row, column=2, value=case.get('status', ''))
            ws.cell(row=row, column=3, value=case.get('priority', ''))
            ws.cell(row=row, column=4, value=case.get('created_at', ''))
            ws.cell(row=row, column=5, value=case.get('resolved_at', ''))
            # Calculate resolution time if both dates exist
            ws.cell(row=row, column=6, value='')  # Placeholder
    
    def _add_audit_summary_sheet(self, ws, data, metadata):
        """Add audit summary"""
        ws['A1'] = metadata.get('title', 'Compliance Audit Report')
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        row = 3
        metrics = [
            ('Total Actions:', data.get('total_actions', 0)),
            ('Blacklist Changes:', data.get('blacklist_changes', 0)),
            ('Screenings Performed:', data.get('screenings_performed', 0)),
            ('Decisions Made:', data.get('decisions_made', 0)),
            ('Compliance Score:', f"{data.get('compliance_score', 0)}%"),
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    def _add_audit_trail_sheet(self, ws, audit_trail):
        """Add audit trail details"""
        headers = ['ID', 'Action Type', 'Entity Name', 'Entity Type', 'Decision', 'User', 'Date', 'IP Address']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        for row, entry in enumerate(audit_trail, 2):
            ws.cell(row=row, column=1, value=entry.get('id', ''))
            ws.cell(row=row, column=2, value=entry.get('action_type', ''))
            ws.cell(row=row, column=3, value=entry.get('entity_name', ''))
            ws.cell(row=row, column=4, value=entry.get('entity_type', ''))
            ws.cell(row=row, column=5, value=entry.get('decision', ''))
            ws.cell(row=row, column=6, value=entry.get('reviewed_by', ''))
            ws.cell(row=row, column=7, value=entry.get('created_at', ''))
            ws.cell(row=row, column=8, value=entry.get('ip_address', ''))


class CSVGenerator:
    """CSV file generator"""
    
    def __init__(self):
        self.reports_dir = "reports"
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_csv(
        self,
        data: List[Dict[str, Any]],
        headers: List[str],
        filename_prefix: str
    ) -> str:
        """
        Generate CSV file
        
        Args:
            data: List of dictionaries
            headers: Column headers
            filename_prefix: Filename prefix
            
        Returns:
            File path
        """
        filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        filepath = os.path.join(self.reports_dir, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.DictWriter(csvfile, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        
        logger.info(f"CSV report generated: {filepath}")
        return filepath


def get_excel_generator() -> ExcelGenerator:
    """Get Excel generator instance"""
    return ExcelGenerator()


def get_csv_generator() -> CSVGenerator:
    """Get CSV generator instance"""
    return CSVGenerator()
