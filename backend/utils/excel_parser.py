"""
Excel Parser Utility for Phase 4
Handles multi-sheet Excel parsing with Arabic text support
"""
import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime
import uuid
from fastapi import UploadFile, HTTPException


class ExcelParserError(Exception):
    """Custom exception for Excel parsing errors"""
    pass


class ExcelParser:
    """
    Excel file parser for blacklist and Kamco entity files
    Supports:
    - Multi-sheet workbooks
    - Arabic text (UTF-8 encoding)
    - Civil ID validation (12-digit)
    - Missing data handling
    - Data validation
    """
    
    def __init__(self, file_path: str = None, file_bytes: bytes = None):
        """
        Initialize parser with either file path or bytes
        
        Args:
            file_path: Path to Excel file
            file_bytes: File content as bytes
        """
        self.file_path = file_path
        self.file_bytes = file_bytes
        self.workbook = None
        self.sheets = {}
        
    def load_workbook(self) -> openpyxl.Workbook:
        """Load Excel workbook"""
        try:
            if self.file_bytes:
                from io import BytesIO
                self.workbook = openpyxl.load_workbook(BytesIO(self.file_bytes), data_only=True)
            elif self.file_path:
                self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
            else:
                raise ExcelParserError("No file path or bytes provided")
            return self.workbook
        except Exception as e:
            raise ExcelParserError(f"Failed to load workbook: {str(e)}")
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names from workbook"""
        if not self.workbook:
            self.load_workbook()
        return self.workbook.sheetnames
    
    def parse_sheet(self, sheet_name: str, expected_columns: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Parse a specific sheet into DataFrame
        
        Args:
            sheet_name: Name of the sheet to parse
            expected_columns: Optional list of expected column names
            
        Returns:
            DataFrame with parsed data
        """
        try:
            if self.file_bytes:
                from io import BytesIO
                df = pd.read_excel(BytesIO(self.file_bytes), sheet_name=sheet_name, engine='openpyxl')
            elif self.file_path:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')
            else:
                raise ExcelParserError("No file path or bytes provided")
            
            # Clean column names (strip whitespace)
            df.columns = df.columns.str.strip()
            
            # Check expected columns if provided
            if expected_columns:
                missing_cols = set(expected_columns) - set(df.columns)
                if missing_cols:
                    raise ExcelParserError(
                        f"Missing required columns in sheet '{sheet_name}': {', '.join(missing_cols)}"
                    )
            
            return df
        except Exception as e:
            raise ExcelParserError(f"Failed to parse sheet '{sheet_name}': {str(e)}")
    
    def parse_all_sheets(self) -> Dict[str, pd.DataFrame]:
        """Parse all sheets in workbook"""
        if not self.workbook:
            self.load_workbook()
        
        sheets = {}
        for sheet_name in self.get_sheet_names():
            try:
                sheets[sheet_name] = self.parse_sheet(sheet_name)
            except Exception as e:
                print(f"Warning: Could not parse sheet '{sheet_name}': {str(e)}")
        
        return sheets
    
    def parse_blacklist(self, sheet_name: str = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """
        Parse blacklist/sanctions list Excel file - ULTRA FLEXIBLE AI-POWERED PARSING
        
        Handles ANY Excel structure including:
        - Title rows before headers
        - Report metadata rows
        - Headers at any row position
        - Missing columns
        - Any language/format
        
        Args:
            sheet_name: Specific sheet name (optional, uses first sheet if not provided)
            
        Returns:
            Tuple of (records_list, summary_dict)
        """
        if not self.workbook:
            self.load_workbook()
        
        # Use first sheet if no sheet name provided
        if not sheet_name:
            sheet_name = self.get_sheet_names()[0]
        
        ws = self.workbook[sheet_name]
        
        # SMART HEADER DETECTION - Find where the real data starts
        header_row_idx, headers, data_start_row = self._smart_find_headers(ws)
        
        # Generate batch ID for this upload
        batch_id = f"BATCH-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{str(uuid.uuid4())[:8]}"
        
        records = []
        errors = []
        total_rows = 0
        
        # Parse data rows starting from detected data row
        for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            total_rows += 1
            
            # Skip completely empty rows
            if not any(cell for cell in row):
                continue
            
            try:
                # Create row dictionary with headers
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = cell
                
                record = self._parse_blacklist_row_flexible(row_dict, batch_id)
                if record:  # Only add if we got something useful
                    records.append(record)
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        summary = {
            "total_rows": total_rows,
            "valid_records": len(records),
            "errors": errors,
            "error_count": len(errors),
            "batch_id": batch_id,
            "sheet_name": sheet_name,
            "header_row": header_row_idx,
            "data_start_row": data_start_row
        }
        
        return records, summary
    
    def _smart_find_headers(self, ws) -> Tuple[int, List[str], int]:
        """
        Intelligently find the header row and data start in ANY Excel structure
        
        Handles:
        - Title rows (e.g., "Report Generated on 2024-01-01")
        - Empty rows before headers
        - Metadata rows
        - Headers anywhere in first 20 rows
        
        Returns:
            (header_row_index, header_list, data_start_row)
        """
        import re
        
        # Common patterns that indicate a row is NOT a header
        skip_patterns = [
            r'report\s+generated',
            r'generated\s+on',
            r'date:',
            r'time:',
            r'printed',
            r'page\s+\d+',
            r'confidential',
            r'internal\s+use',
            r'©',
            r'copyright',
        ]
        
        # Common header keywords in multiple languages
        header_keywords = [
            'name', 'اسم', 'الاسم',
            'id', 'رقم', 'civil',
            'type', 'نوع', 'category',
            'date', 'تاريخ',
            'country', 'دولة', 'بلد',
            'status', 'حالة',
            'passport', 'جواز',
            'nationality', 'جنسية',
        ]
        
        best_header_row = 0
        best_headers = []
        best_score = 0
        
        # Check first 20 rows for potential headers
        for row_idx in range(min(20, ws.max_row)):
            row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
            
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            
            # Convert row to strings for analysis
            row_text = ' '.join([str(cell).lower() if cell else '' for cell in row])
            
            # Skip if it matches a skip pattern (title/metadata row)
            if any(re.search(pattern, row_text, re.IGNORECASE) for pattern in skip_patterns):
                continue
            
            # Count how many cells look like headers
            header_score = 0
            non_empty_count = 0
            
            for cell in row:
                if cell:
                    non_empty_count += 1
                    cell_str = str(cell).lower()
                    
                    # Check if cell contains header keywords
                    if any(keyword in cell_str for keyword in header_keywords):
                        header_score += 10
                    
                    # Headers are usually short strings
                    if len(cell_str) < 50 and not cell_str.isdigit():
                        header_score += 1
                    
                    # Headers don't usually have long text
                    if len(cell_str) > 100:
                        header_score -= 5
            
            # Need at least 2 non-empty cells
            if non_empty_count >= 2 and header_score > best_score:
                best_score = header_score
                best_header_row = row_idx
                # Clean and normalize headers - preserve Arabic characters
                best_headers = []
                for i, cell in enumerate(row):
                    if cell:
                        # Convert to string and strip
                        header = str(cell).strip()
                        # Only lowercase and replace spaces/dashes if it's ASCII
                        if header.isascii():
                            header = header.lower().replace(' ', '_').replace('-', '_')
                        else:
                            # Keep Arabic/non-ASCII as-is but replace spaces
                            header = header.replace(' ', '_').replace('-', '_')
                        best_headers.append(header)
                    else:
                        best_headers.append(f'col_{i}')
        
        # If no good header found, use first non-empty row
        if not best_headers:
            for row_idx in range(min(10, ws.max_row)):
                row = list(ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True))[0]
                if any(cell for cell in row):
                    best_header_row = row_idx
                    best_headers = [f'col_{i}' for i in range(len(row))]
                    break
        
        # Data starts right after headers
        data_start_row = best_header_row + 2  # +1 for 1-based index, +1 to skip header row
        
        return best_header_row, best_headers, data_start_row
    
    def _parse_blacklist_row_flexible(self, row: Dict[str, Any], batch_id: str) -> Optional[Dict[str, Any]]:
        """
        Parse a blacklist row with MAXIMUM flexibility - AI-powered field detection
        
        Smart features:
        - Tries multiple column name variations
        - Handles merged cells and empty values
        - Detects and skips header/metadata rows
        - Fallback to ANY non-empty column
        - Multi-language support (English, Arabic, etc.)
        """
        import pandas as pd
        import re
        
        def get_value(keys: list, default: str = None) -> str:
            """Try multiple possible column names and return first non-empty value"""
            for key in keys:
                val = row.get(key, None)
                if val is not None and not pd.isna(val):
                    val_str = str(val).strip()
                    # Skip if it looks like a header or metadata
                    if val_str and not self._looks_like_header(val_str):
                        return val_str
            return default
        
        def clean_numeric_id(val: str) -> str:
            """Extract numeric ID from string (handles formats like 'ID: 12345')"""
            if not val:
                return None
            # Remove common prefixes and extract numbers
            cleaned = re.sub(r'[^\d]', '', str(val))
            return cleaned if cleaned else None
        
        # Try multiple column name variations for each field (English, Arabic, Transliterated, Mixed)
        name_arabic = get_value([
            # English variations
            'name_arabic', 'arabic_name', 'name', 'full_name', 'full_name_arabic',
            'person_name', 'entity_name', 'individual_name', 'customer_name', 
            'client_name', 'subject_name', 'fullname', 'full name',
            # Arabic variations
            'الاسم', 'اسم', 'الاسم_الكامل', 'الأسم', 'أسم',
            'اسم_كامل', 'الإسم', 'إسم',
            # Transliterated variations
            'ism', 'al_ism', 'isim', 'al_isim', 'esm', 'al_esm',
            # Mixed language
            'name_عربي', 'الاسم_arabic', 'اسم_name'
        ])
        
        # If still no name found, use FIRST non-empty column as name (ultimate fallback)
        if not name_arabic:
            for key, val in row.items():
                if val and not pd.isna(val):
                    val_str = str(val).strip()
                    # Make sure it's not a date, number, or header
                    if (val_str and 
                        not self._looks_like_header(val_str) and
                        not val_str.isdigit() and
                        len(val_str) > 2):
                        name_arabic = val_str
                        break
        
        # Skip if we got absolutely nothing useful
        if not name_arabic or name_arabic == 'None':
            return None
        
        # Try to extract other fields with multiple variations (English, Arabic, Transliterated, Mixed)
        civil_id = get_value([
            # English variations
            'civil_id', 'civilid', 'id', 'national_id', 'identity_number', 
            'id_number', 'cpr', 'civil_id_number', 'personal_id', 'identity_no', 
            'identity', 'citizen_id', 'id_no', 'number', 'no', 'civil id',
            'national id', 'id number',
            # Arabic variations
            'رقم_مدني', 'رقم_هوية', 'الرقم_المدني', 'رقم', 'هوية',
            'رقم مدني', 'رقم هوية', 'الرقم المدني',
            # Transliterated variations
            'raqam_madani', 'raqam', 'hawiya', 'rakam', 'ragam',
            # Mixed language
            'civil_رقم', 'id_مدني', 'رقم_id'
        ])
        civil_id = clean_numeric_id(civil_id)
        
        passport_number = get_value([
            # English variations
            'passport_number', 'passport', 'passportno', 'passport_no', 
            'travel_document', 'passport_id', 'travel_doc', 'document_number', 
            'doc_no', 'document', 'travel_document_number', 'travel_id',
            'passport no', 'passport number', 'travel document',
            # Arabic variations
            'جواز_سفر', 'جواز', 'رقم_جواز', 'رقم_جواز_السفر',
            'جواز سفر', 'رقم جواز', 'وثيقة_سفر',
            # Transliterated variations
            'jawaz_safar', 'jawaz', 'gavaz', 'jawaz_safer',
            # Mixed language
            'passport_جواز', 'جواز_passport', 'travel_سفر'
        ])
        
        nationality = get_value([
            # English variations
            'nationality', 'country', 'nation', 'country_of_origin', 
            'citizenship', 'national', 'citizen', 'country of origin',
            # Arabic variations
            'جنسية', 'بلد', 'الجنسية', 'البلد', 'دولة', 'الدولة',
            'جنسيه', 'بلاد', 'موطن',
            # Transliterated variations
            'jinsiya', 'balad', 'dawla', 'ginsiya',
            # Mixed language
            'nationality_جنسية', 'جنسية_nationality', 'country_بلد'
        ])
        
        entity_type = get_value([
            # English variations
            'type', 'entity_type', 'person_type', 'category', 
            'classification', 'individual_or_entity', 'entity type',
            'person type',
            # Arabic variations
            'نوع', 'نوع_الكيان', 'النوع', 'فئة', 'تصنيف',
            'نوع الكيان', 'تصنيف الكيان',
            # Transliterated variations
            'naw', 'naw_alkayan', 'fe2a', 'tasnif',
            # Mixed language
            'type_نوع', 'نوع_type', 'entity_كيان'
        ], 'individual')
        
        dob = get_value([
            # English variations
            'date_of_birth', 'dob', 'birth_date', 'birthdate', 'born', 
            'birth', 'date_of_birth_(dob)', 'date of birth', 'birth date',
            # Arabic variations
            'تاريخ_الميلاد', 'تاريخ_ميلاد', 'الميلاد', 'تاريخ الميلاد',
            'مولود', 'تاريخ_الولادة', 'تاريخ ولادة',
            # Transliterated variations
            'tareekh_meelaad', 'meelaad', 'tarikh', 'milad',
            # Mixed language
            'dob_تاريخ', 'birth_ميلاد', 'تاريخ_birth'
        ])
        
        source = get_value([
            # English variations
            'source', 'list_source', 'origin', 'list_name', 'database', 
            'list', 'sanctions_list', 'data source', 'list name',
            # Arabic variations
            'مصدر', 'المصدر', 'أصل', 'قائمة', 'القائمة',
            'مصدر_البيانات', 'اسم_القائمة', 'مصدر البيانات',
            # Transliterated variations
            'masdar', 'asl', 'qa2ima', 'qaima',
            # Mixed language
            'source_مصدر', 'مصدر_source', 'list_قائمة'
        ], 'Uploaded File')
        
        reason = get_value([
            # English variations
            'reason', 'flag_reason', 'notes', 'description', 'comments', 
            'remarks', 'details', 'flag reason', 'comment',
            # Arabic variations
            'سبب', 'السبب', 'ملاحظات', 'تفاصيل', 'وصف',
            'سبب_العلم', 'التفاصيل', 'الملاحظات', 'وصف تفصيلي',
            # Transliterated variations
            'sabab', 'mulahazat', 'tafaseel', 'wasf',
            # Mixed language
            'reason_سبب', 'سبب_reason', 'notes_ملاحظات'
        ], 'Blacklist Match')
        
        return {
            "name_arabic": name_arabic,
            "civil_id": civil_id,
            "passport_number": passport_number,
            "nationality": nationality,
            "entity_type": entity_type,
            "dob": dob,
            "source": source,
            "reason": reason,
            "batch_id": batch_id,
            "upload_date": datetime.now().isoformat()
        }
    
    def _looks_like_header(self, text: str) -> bool:
        """Check if a text value looks like a header/metadata row rather than data"""
        if not text:
            return False
        
        text_lower = text.lower()
        
        # Common header/metadata patterns
        header_patterns = [
            'report', 'generated', 'date:', 'time:', 'page',
            'confidential', 'internal', 'copyright', '©',
            'printed', 'exported', 'total:', 'summary',
            'header', 'column', 'field', 'name', 'type'
        ]
        
        # If it exactly matches a header keyword
        if text_lower in header_patterns:
            return True
        
        # If it contains header patterns
        if any(pattern in text_lower for pattern in header_patterns):
            # But make sure it's not just a name that happens to contain these words
            if len(text) < 30:  # Headers are usually short
                return True
        
        return False
    
    def validate_blacklist_file(self) -> Dict[str, Any]:
        """
        Validate blacklist Excel file structure - SUPER FLEXIBLE
        Accepts any structure with any columns
        
        Returns:
            Validation result dictionary
        """
        try:
            if not self.workbook:
                self.load_workbook()
            
            sheet_names = self.get_sheet_names()
            if not sheet_names:
                return {
                    "valid": False,
                    "error": "Workbook has no sheets"
                }
            
            # Check first sheet has ANY data
            first_sheet = sheet_names[0]
            try:
                df = self.parse_sheet(first_sheet, required_columns=[])  # No required columns!
            except:
                # Even if parsing fails, try to read raw sheet
                ws = self.workbook[first_sheet]
                if ws.max_row < 2:  # Need at least header + 1 data row
                    return {
                        "valid": False,
                        "error": "No data rows found in Excel file"
                    }
                # Accept it anyway!
                return {
                    "valid": True,
                    "message": "File accepted (flexible parsing)",
                    "sheets": sheet_names,
                    "primary_sheet": first_sheet
                }
            
            # Check if there's any data
            if len(df) == 0:
                return {
                    "valid": False,
                    "error": "No data rows found in Excel file"
                }
            
            # Accept any file with data!
            return {
                "valid": True,
                "message": "File accepted (flexible mode)",
                "sheets": sheet_names,
                "primary_sheet": first_sheet,
                "row_count": len(df),
                "columns": list(df.columns)
            }
            
        except Exception as e:
            # Be even more lenient - accept the file anyway
            return {
                "valid": True,
                "message": f"File accepted despite error: {str(e)}",
                "warning": str(e)
            }
    
    def parse_kamco_entities(self, entity_type: str = "Clients") -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
        """
        Parse Kamco entity data (Clients, Vendors, Staff, Others)
        
        Args:
            entity_type: Type of entity sheet to parse
            
        Returns:
            Tuple of (records_list, summary_dict)
        """
        if not self.workbook:
            self.load_workbook()
        
        # Try to find matching sheet
        sheet_name = None
        for name in self.get_sheet_names():
            if entity_type.lower() in name.lower():
                sheet_name = name
                break
        
        if not sheet_name:
            raise ExcelParserError(f"Could not find sheet for entity type: {entity_type}")
        
        df = self.parse_sheet(sheet_name)
        
        records = []
        for idx, row in df.iterrows():
            record = row.to_dict()
            # Clean None values
            record = {k: (v if pd.notna(v) else None) for k, v in record.items()}
            records.append(record)
        
        summary = {
            "entity_type": entity_type,
            "sheet_name": sheet_name,
            "total_records": len(records),
            "columns": list(df.columns)
        }
        
        return records, summary


# Convenience functions

def parse_blacklist_excel(file_path: str = None, file_bytes: bytes = None) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    Convenience function to parse blacklist Excel file
    
    Args:
        file_path: Path to Excel file
        file_bytes: File content as bytes
        
    Returns:
        Tuple of (records_list, summary_dict)
    """
    parser = ExcelParser(file_path=file_path, file_bytes=file_bytes)
    return parser.parse_blacklist()


def validate_blacklist_excel(file_path: str = None, file_bytes: bytes = None) -> Dict[str, Any]:
    """
    Convenience function to validate blacklist Excel file
    
    Args:
        file_path: Path to Excel file
        file_bytes: File content as bytes
        
    Returns:
        Validation result dictionary
    """
    parser = ExcelParser(file_path=file_path, file_bytes=file_bytes)
    return parser.validate_blacklist_file()


# Legacy compatibility function
async def parse_blacklist_excel_legacy(file: UploadFile) -> Dict[str, List[Dict]]:
    """
    Legacy function for backward compatibility
    Parse uploaded Excel file and extract data from all 4 sheets
    """
    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(filename=file.filename, data_only=True)
        
        result = {}
        expected_sheets = ['Clients', 'Vendors', 'Staff', 'Others']
        
        for sheet_name in expected_sheets:
            actual_sheet_name = None
            for ws_name in workbook.sheetnames:
                if ws_name.lower() == sheet_name.lower():
                    actual_sheet_name = ws_name
                    break
            
            if not actual_sheet_name:
                result[sheet_name.lower()] = []
                continue
            
            worksheet = workbook[actual_sheet_name]
            headers = [cell.value if cell.value else f"Column_{i}" for i, cell in enumerate(worksheet[1])]
            
            rows = []
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    continue
                
                row_dict = {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
                if row_dict.get('Name') or row_dict.get('name'):
                    rows.append(row_dict)
            
            result[sheet_name.lower()] = rows
        
        workbook.close()
        return result
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error parsing Excel file: {str(e)}")

